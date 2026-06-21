"""Parser for the "tzzb" portfolio-tracker xlsx export (e.g. Huabao Securities).

tzzb = 同花顺投资账本 (Tonghuashun Investment Ledger), the source app.

The workbook has 持仓数据 / 已清仓 / 交易记录 sheets; only 交易记录 (the
transaction log) is parsed. The other sheets are derived state, not transactions.
"""

import datetime as dt
from collections import Counter

from openpyxl import load_workbook

import schema
import symbols

TXN_SHEET = "交易记录"

# Expected 交易记录 header -> canonical field (or None to ignore).
HEADER_MAP = {
    "成交日期": "date",
    "成交时间": "time",
    "代码": "symbol",
    "名称": "name",
    "交易类别": "type_raw",
    "成交数量": "quantity",
    "成交价格": "price",
    "发生金额": "cash_flow",
    "成交金额": "gross_amount",
    "费用": "fee",
    "备注": "note",
}

NAME = "tzzb"

# tzzb 交易类别 (transaction type) -> normalized action (schema.ACTIONS).
# This vocabulary is broker-specific; other brokers define their own.
TYPE_TO_ACTION = {
    "买入": "BUY",
    "融资买入": "BUY",
    "卖出": "SELL",
    "融券卖出": "SELL",
    "股份转入": "TRANSFER_IN",
    "转托转入": "BUY",        # transfer-custody in (shares moved from another custodian)
    "转债转入": "BUY",        # convertible-bond allotment: booked as a buy
    "股份转出": "TRANSFER_OUT",
    "转债转出": "TRANSFER_OUT",
    "卖券还款": "SELL",       # sell securities to repay margin loan
    "红利入账": "DIVIDEND",
    "股息": "DIVIDEND",
    "除权除息": "DIVIDEND",   # ex-dividend cash payout
    "利息": "INTEREST",
    "银证转入": "DEPOSIT",
    "入金": "DEPOSIT",
    "收入": "DEPOSIT",        # cash income / top-up (no security)
    "银证转出": "WITHDRAWAL",
    "出金": "WITHDRAWAL",
    "拆股": "TRANSFER_IN",    # bonus shares: a free share transfer-in
    "组合费用": "FEE",        # HKEX daily levy charged on the portfolio
    "股息个税征收": "TAX",     # individual income tax on dividends
}

# tzzb types that involve margin financing -> canonical subtype MARGIN.
MARGIN_TYPES = {"融资买入", "融券卖出", "卖券还款", "融券", "融券购回"}


def action_for(type_raw: str, code: str = "") -> str:
    """Map a tzzb type string to a normalized action (OTHER if unknown).

    `融券` / `融券购回` are context-dependent and need the security `code`:
      - government-bond reverse repo (e.g. GC001/204001, R-001/131810):
        `融券` is lending cash out (BUY), `融券购回` is getting it back (SELL).
      - otherwise it is true short selling: `融券` is SELL, `融券购回` is BUY.
    """
    t = (type_raw or "").strip()
    if t in ("融券", "融券购回"):
        if symbols.is_reverse_repo(code):
            return "BUY" if t == "融券" else "SELL"
        return "SELL" if t == "融券" else "BUY"
    return TYPE_TO_ACTION.get(t, "OTHER")


def matches(path: str) -> bool:
    """True if this looks like a tzzb workbook (has a 交易记录 sheet)."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        return TXN_SHEET in wb.sheetnames
    finally:
        wb.close()


def _norm_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _norm_time(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.time)):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def _num(value):
    """Return a number unchanged, "" for blanks; pass other strings through."""
    if value is None or value == "":
        return ""
    return value


def _is_zero(value) -> bool:
    """True if the cell is blank or numerically zero."""
    if value is None or value == "":
        return True
    try:
        return abs(float(value)) < 1e-9
    except (TypeError, ValueError):
        return False


def parse(path: str, account: str, source_file: str):
    """Yield canonical row dicts from the 交易记录 sheet."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[TXN_SHEET]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return
        # Column index -> canonical field, based on the actual header order.
        idx_field = {}
        for i, col in enumerate(header):
            name = str(col).strip() if col is not None else None
            field = HEADER_MAP.get(name) if name else None
            if field:
                idx_field[i] = field

        # Counts identical identifying tuples seen so far (file order) so
        # genuine same-instant repeats get distinct, stable txn_ids.
        occ = Counter()
        for raw in rows:
            rec = {f: "" for f in schema.COLUMNS}
            for i, field in idx_field.items():
                if i < len(raw):
                    rec[field] = raw[i]

            # Skip blank/summary rows (no date and no symbol).
            date = _norm_date(rec["date"])
            symbol = "" if rec["symbol"] is None else str(rec["symbol"]).strip()
            if not date and not symbol:
                continue
            if symbol in ("汇总", "合计", "总计"):
                continue
            # Skip no-op rows: zero quantity AND zero cash flow (placeholder /
            # cancelled entries that affect neither position nor cash).
            if _is_zero(rec["quantity"]) and _is_zero(rec["cash_flow"]):
                continue

            name = "" if rec["name"] is None else str(rec["name"]).strip()
            type_raw = rec["type_raw"]
            type_raw = "" if type_raw is None else str(type_raw).strip()
            note = "" if rec["note"] is None else str(rec["note"]).strip()

            rec["account"] = account
            rec["date"] = date
            rec["time"] = _norm_time(rec["time"])
            rec["symbol"] = symbol
            rec["name"] = name
            rec["type_raw"] = type_raw
            rec["action"] = action_for(type_raw, symbol)
            # Pure cash movements carry the canonical $CASH symbol (currency is
            # still taken from the raw code below, so an HKD levy stays HKD).
            if rec["action"] in schema.CASH_ACTIONS:
                rec["symbol"] = schema.CASH_SYMBOL
            rec["subtype"] = "MARGIN" if type_raw in MARGIN_TYPES else ""
            rec["quantity"] = _num(rec["quantity"])
            rec["price"] = _num(rec["price"])
            rec["cash_flow"] = _num(rec["cash_flow"])
            rec["gross_amount"] = _num(rec["gross_amount"])
            rec["fee"] = _num(rec["fee"])
            rec["currency"] = symbols.currency_of(symbol)
            rec["note"] = note
            rec["source_file"] = source_file

            key = (date, rec["time"], symbol, type_raw,
                   rec["quantity"], rec["price"], rec["cash_flow"])
            rec["txn_id"] = schema.make_txn_id(rec, occ[key])
            occ[key] += 1
            yield rec
    finally:
        wb.close()
